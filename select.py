# -*- coding: utf-8 -*-
"""
Created on Tue Oct 30 19:03:22 2018

@author: Neoooli
"""
import numpy as np
import os
import glob
from gdaldiy import *
import openpyxl
"""
 the file should be organized as:
 ./train
        /cloud
             /10mclips/imagename/patch_Id.tif
             /20mclips/imagename/patch_Id.tif
             /60mclips/imagename/patch_Id.tif
  ./train
        /clear
             /10mclips/imagename/patch_Id.tif
             /20mclips/imagename/patch_Id.tif
             /60mclips/imagename/patch_Id.tif
 ./test
        /cloud
             /10mclips/imagename/patch_Id.tif
             /20mclips/imagename/patch_Id.tif
             /60mclips/imagename/patch_Id.tif
 ./test
        /clear
             /10mclips/imagename/patch_Id.tif
             /20mclips/imagename/patch_Id.tif
             /60mclips/imagename/patch_Id.tif
"""
commendir='F:/WHU/WHUS2-CR/composite/'#source dir
filedirs=["cloud/","clear/"]
targetnames=['10mclips','20mclips','60mclips']
filetypes=['train','test']
  
def selectfromxls(xlpath,filedir,targetname):
    wb=openpyxl.load_workbook(xlpath)
    for filetype in filetypes:
        ws=wb[filetype]#fileno for train data.
        for i in range(1,ws.max_row+1):
            name = str(ws.cell(row=i,column=1).value)
            pairpath=commendir+filedir+targetname+name
            savedir = commendir+"selected/"+filedir+targetname+'/'+name.split('/')[1]
            if not os.path.exists(pairpath): 
                continue
            if not os.path.exists(savedir): 
                os.makedirs(savedir)                                
            savepath=savedir+'/'+name.split('/')[-1]
            if not os.path.exists(savepath):
                selectedimg=imgread(pairpath)#read file function. You can replace this function for reading images.
                imgwrite(savepath,selectedimg)##save file function.You can replace this function for saving images.
for filedir in filedirs:
    for targetname in targetnames:
        selectfromxls("./filenos.xlsx",filedir,targetname)

